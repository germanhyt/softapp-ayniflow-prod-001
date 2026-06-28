import unicodedata
from io import BytesIO

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook

from app.modules.finance.domain.models import Transaction


def _pdf_safe(value: object, *, max_len: int | None = None) -> str:
    text = "" if value is None else str(value)
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    if max_len is not None:
        return ascii_text[:max_len]
    return ascii_text


class FinanceExportService:
    def build_excel(self, transactions: list[Transaction], summary: dict[str, str]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transacciones"

        sheet.append(["AyniFlow - Reporte Financiero"])
        sheet.append(["Ingresos", summary["total_income"]])
        sheet.append(["Egresos", summary["total_expense"]])
        sheet.append(["Balance", summary["balance"]])
        sheet.append([])

        headers = [
            "Fecha",
            "Hora",
            "Movimiento",
            "Concepto",
            "Banco",
            "Tipo",
            "Destinatario",
            "N° Operación",
            "Monto",
            "Categoría",
        ]
        sheet.append(headers)

        for tx in transactions:
            sheet.append(
                [
                    tx.transaction_date.isoformat(),
                    tx.transaction_time.isoformat(timespec="minutes") if tx.transaction_time else "",
                    tx.movement_type.value,
                    tx.concept,
                    tx.bank,
                    tx.payment_type,
                    tx.recipient or "",
                    tx.operation_number or "",
                    float(tx.amount),
                    tx.category or "",
                ]
            )

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def build_pdf(self, transactions: list[Transaction], summary: dict[str, str]) -> bytes:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "AyniFlow - Reporte Financiero", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(
            0,
            8,
            f"Ingresos: {_pdf_safe(summary['total_income'])}",
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )
        pdf.cell(
            0,
            8,
            f"Egresos: {_pdf_safe(summary['total_expense'])}",
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )
        pdf.cell(
            0,
            8,
            f"Balance: {_pdf_safe(summary['balance'])}",
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )
        pdf.ln(4)

        pdf.set_font("Helvetica", "B", 9)
        columns = ["Fecha", "Mov.", "Concepto", "Banco", "Monto"]
        widths = [25, 18, 65, 25, 25]
        for idx, title in enumerate(columns):
            pdf.cell(widths[idx], 8, title, border=1)
        pdf.ln()

        pdf.set_font("Helvetica", size=8)
        for tx in transactions[:80]:
            values = [
                tx.transaction_date.isoformat(),
                _pdf_safe(tx.movement_type.value, max_len=6),
                _pdf_safe(tx.concept, max_len=35),
                _pdf_safe(tx.bank, max_len=12),
                _pdf_safe(tx.amount),
            ]
            for idx, value in enumerate(values):
                pdf.cell(widths[idx], 7, value, border=1)
            pdf.ln()

        if len(transactions) > 80:
            pdf.ln(2)
            pdf.set_font("Helvetica", size=8)
            pdf.cell(
                0,
                6,
                _pdf_safe(f"Mostrando 80 de {len(transactions)} transacciones."),
                new_x=XPos.LMARGIN,
                new_y=YPos.NEXT,
            )

        output = pdf.output()
        return bytes(output) if isinstance(output, bytearray) else output
